# -*- coding: utf-8 -*-
import json
from dateutil import parser
from openpyxl import load_workbook
from anyway.parsers.utils import batch_iterator
from anyway.backend_constants import BE_CONST
from anyway.models import AccidentMarker
from anyway.app_and_db import db


def _iter_rows(filename):
    workbook = load_workbook(filename, read_only=True)
    sheet = workbook["Worksheet1"]
    rows = sheet.rows
    first_row = next(rows)
    headers = [
        "מזהה",
        "תאריך דיווח",
        "סטטוס",
        "סוג עבירה",
        "סוג רכב",
        "סוג לוחית רישוי",
        "נ״צ ערוך",
    ]
    assert [cell.value for cell in first_row] == headers
    for row in rows:
        id_ = int(row[0].value)
        provider_and_id_ = int(str(BE_CONST.RSA_PROVIDER_CODE) + str(id_))

        violation = row[3].value
        vehicle_type = row[4].value
        coordinates = row[6].value
        if coordinates == "," or coordinates == "0.0,0.0":
            continue
        rsa_license_plate = row[5].value
        video_link = None
        timestamp = parser.parse(row[1].value, dayfirst=True)
        if not violation:
            continue

        vehicle_type = vehicle_type or ""
        coordinates = coordinates.split(",")
        latitude, longitude = float(coordinates[0]), float(coordinates[1])
        description = {
            "VIOLATION_TYPE": violation,
            "VEHICLE_TYPE": vehicle_type,
            "RSA_LICENSE_PLATE": rsa_license_plate,
        }

        yield {
            "id": id_,
            "provider_and_id": provider_and_id_,
            "latitude": latitude,
            "longitude": longitude,
            "created": timestamp,
            "provider_code": BE_CONST.RSA_PROVIDER_CODE,
            "accident_severity": 0,
            "title": "שומרי הדרך",
            "description": json.dumps(description),
            "location_accuracy": 1,
            "type": BE_CONST.MARKER_TYPE_ACCIDENT,
            "video_link": video_link,
            "vehicle_type_rsa": vehicle_type,
            "violation_type_rsa": violation,
            "rsa_license_plate": rsa_license_plate,
            "accident_year": timestamp.year,
        }


def parse(filename):
    db.session.execute(f"DELETE from markers where provider_code = {BE_CONST.RSA_PROVIDER_CODE}")
    for batch in batch_iterator(_iter_rows(filename), batch_size=50000):
        db.session.bulk_insert_mappings(AccidentMarker, batch)
        db.session.commit()

    """
    Fills empty geometry object according to coordinates in database
    """
    db.session.execute(
        "UPDATE markers SET geom = ST_SetSRID(ST_MakePoint(longitude,latitude),4326)\
                           WHERE geom IS NULL;"
    )
    db.session.commit()
