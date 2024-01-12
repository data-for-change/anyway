# -*- coding: utf-8 -*-
import sys
from typing import Dict, Tuple, Iterator
import logging
from openpyxl import load_workbook
from anyway.app_and_db import db
from anyway.models import SuburbanJunction, RoadJunctionKM


SUBURBAN_JUNCTION = "suburban_junction"
ACCIDENTS = "accidents"
CITIES = "cities"
STREETS = "streets"
ROADS = "roads"
URBAN_INTERSECTION = "urban_intersection"
NON_URBAN_INTERSECTION = "non_urban_intersection"
NON_URBAN_INTERSECTION_HEBREW = "non_urban_intersection_hebrew"
DICTIONARY = "dictionary"
INVOLVED = "involved"
VEHICLES = "vehicles"
ID = "id"
NAME = "name"
KM = "km"
ROAD1 = "road1"
suburban_junctions_dict: Dict[int, dict] = {}
# (road, junction) -> km
road_junction_km_dict: Dict[Tuple[int, int], int] = {}


def parse(filename):
    read_from_file(filename)
    import_suburban_junctions_into_db()
    import_road_junction_km_into_db()


def read_from_file(filename: str):
    for j in _iter_rows(filename):
        add_suburban_junction(j)
        add_road_junction_km(j)


def _iter_rows(filename) -> Iterator[dict]:
    workbook = load_workbook(filename, read_only=True)
    sheet = workbook["מילון צמתים לא עירוניים"]
    rows = sheet.rows
    first_row = next(rows)
    headers = [
        "ZOMET",
        "SUG_DEREH",
        "REHOV1_KVISH1",
        "REHOV2_KVISH2",
        "KM",
        "IKS",
        "IGREK",
        "IDF",
        "SHEM_ZOMET",
        "SUG_ZOMET",
        "KVISH_RASHI",
        "KM_RASHI",
        "SHNAT_ZOMET_SGIRA",
        "MAHOZ",
        "NAFA",
        "EZOR_TIVI",
        "METROPOLIN",
        "MAAMAD_MINIZIPALI",
        "EZOR_STAT",
    ]
    assert [cell.value for cell in first_row] == headers, "File does not have expected headers"
    for row in rows:
        # In order to ignore empty lines
        if not row[0].value:
            continue
        yield {ID: row[0].value, NAME: row[8].value, ROAD1: row[2].value, KM: row[4].value}


def add_road_junction_km(junction: dict):
    road_junction_km_dict[(junction[ROAD1], junction[ID])] = junction[KM] / 10


def import_suburban_junctions_into_db():
    items = [
        {
            "non_urban_intersection": k,
            NON_URBAN_INTERSECTION_HEBREW: fix_name_len(v[NON_URBAN_INTERSECTION_HEBREW]),
            ROADS: v[ROADS],
        }
        for k, v in suburban_junctions_dict.items()
    ]
    logging.debug(f"Writing to db: {len(items)} suburban junctions")
    db.session.query(SuburbanJunction).delete()
    db.session.bulk_insert_mappings(SuburbanJunction, items)
    db.session.commit()
    logging.debug(f"Done writing SuburbanJunction.")


def import_road_junction_km_into_db():
    items = [
        {"road": k[0], "non_urban_intersection": k[1], "km": v}
        for k, v in road_junction_km_dict.items()
    ]
    logging.debug(f"Writing to db: {len(items)} road junction km rows")
    db.session.query(RoadJunctionKM).delete()
    db.session.bulk_insert_mappings(RoadJunctionKM, items)
    db.session.commit()
    logging.debug(f"Done writing RoadJunctionKM.")


def fix_name_len(name: str) -> str:
    if not isinstance(name, str):
        return name
    if len(name) > SuburbanJunction.MAX_NAME_LEN:
        logging.error(
            f"Suburban_junction name too long ({len(name)}>"
            f"{SuburbanJunction.MAX_NAME_LEN}):{name}."
        )
    return name[: SuburbanJunction.MAX_NAME_LEN]


def add_suburban_junction(junction: dict):
    j_id = junction[ID]
    j_name = junction[NAME]
    road1 = junction[ROAD1]
    if j_id in suburban_junctions_dict:
        existing_junction = suburban_junctions_dict[j_id]
        existing_junction[ROADS].add(road1)
    else:
        suburban_junctions_dict[j_id] = {
            NON_URBAN_INTERSECTION_HEBREW: j_name,
            ROADS: {road1},
        }


if __name__ == "__main__":
    parse(sys.argv[1])
